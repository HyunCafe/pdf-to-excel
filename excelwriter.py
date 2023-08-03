from re import sub
import openpyxl
import regex
from openpyxl.styles import Font
import paths

path = paths.excelPath
wb_obj = openpyxl.load_workbook(path) 
wb = openpyxl.Workbook() 
sheet = wb.active 

font = Font(size=15)  # font size

# Column Numbers
group_code_Column = 1
group_description_Column = 2
desc_Column = 3
age_Column = 4
condition_Column = 5
qty_Column = 6
item_amount_Column = 7
unit_Column = 8
reported_cost_Column = 9
unit_cost_Column = 10
coverage_Column = 11
sales_tax_Column = 12
rcv_Column = 13
life_Column = 14
depreciation_type_Column = 15
depreciation_amount_Column = 16
recoverable_Column = 17
acv_Column = 18
tax_Column = 19
replace_Column = 20
cat_Column = 21
sel_Column = 22
owner_Column = 23
original_vendor_Column = 24
date_Column = 25
note_1_Column = 26

# title cell positions
group_code = sheet.cell(row=1, column=group_code_Column)  # group code column title
group_description = sheet.cell(row=1, column=group_description_Column)  # group description column title
desc = sheet.cell(row=1, column=desc_Column)  # desc column title
age = sheet.cell(row=1, column=age_Column)  # age column title
condition = sheet.cell(row=1, column=condition_Column)  # condition column title
qty = sheet.cell(row=1, column=qty_Column)  # qty column title
item_amount = sheet.cell(row=1, column=item_amount_Column)  # item amount column title
unit = sheet.cell(row=1, column=unit_Column)  # unit column title
reported_cost = sheet.cell(row=1, column=reported_cost_Column)  # reported cost column title
unit_cost = sheet.cell(row=1, column=unit_cost_Column)  # unit cost column title
coverage = sheet.cell(row=1, column=coverage_Column)  # coverage column title
sales_tax = sheet.cell(row=1, column=sales_tax_Column)  # sales tax column title
rcv = sheet.cell(row=1, column=rcv_Column)  # rcv column title
life = sheet.cell(row=1, column=life_Column)  # life column title
depreciation_type = sheet.cell(row=1, column=depreciation_type_Column)  # depreciation type column title
depreciation_amount = sheet.cell(row=1, column=depreciation_amount_Column)  # depreciation amount column title
recoverable = sheet.cell(row=1, column=recoverable_Column)  # recoverable column title
acv = sheet.cell(row=1, column=acv_Column)  # acv column title
tax = sheet.cell(row=1, column=tax_Column)  # tax column title
replace = sheet.cell(row=1, column=replace_Column)  # replace column title
cat = sheet.cell(row=1, column=cat_Column)  # cat column title
sel = sheet.cell(row=1, column=sel_Column)  # sel column title
owner = sheet.cell(row=1, column=owner_Column)  # owner column title
original_vendor = sheet.cell(row=1, column=original_vendor_Column)  # original vendor column title
date = sheet.cell(row=1, column=date_Column)  # date column title
note_1 = sheet.cell(row=1, column=note_1_Column)  # note 1 column title

# title names
group_code.value = "Group Code"
group_description.value = "Group Description"
desc.value = "Desc"
age.value = "Age"
condition.value = "Condition"
qty.value = "Qty"
item_amount.value = "Item Amount"
unit.value = "Unit"
reported_cost.value = "Reported Cost"
unit_cost.value = "Unit Cost"
coverage.value = "Coverage"
sales_tax.value = "Sales Tax"
rcv.value = "RCV"
life.value = "Life"
depreciation_type.value = "Depreciation Type"
depreciation_amount.value = "Depreciation Amount"
recoverable.value = "Recoverable"
acv.value = "ACV"
tax.value = "Tax"
replace.value = "Replace"
cat.value = "Cat"
sel.value = "Sel"
owner.value = "Owner"
original_vendor.value = "Original Vendor"
date.value = "Date"
note_1.value = "Note 1"

# sets title fonts
group_code.font = font
group_description.font = font
desc.font = font
age.font = font
condition.font = font
qty.font = font
item_amount.font = font
unit.font = font
reported_cost.font = font
unit_cost.font = font
coverage.font = font
sales_tax.font = font
rcv.font = font
life.font = font
depreciation_type.font = font
depreciation_amount.font = font
recoverable.font = font
acv.font = font
tax.font = font
replace.font = font
cat.font = font
sel.font = font
owner.font = font
original_vendor.font = font
date.font = font
note_1.font = font

def start():
    k = 0
    new_row = 2
    # inserting values to excel
while(new_row < len(excluded) + 2):
    insert_to_excel(excluded[k], new_row, excluded_Column)
    insert_to_excel(group_code[k], new_row, group_code_Column)
    insert_to_excel(group_description[k], new_row, group_description_Column)
    insert_to_excel(desc[k], new_row, desc_Column)
    insert_to_excel(age[k], new_row, age_Column)
    insert_to_excel(condition[k], new_row, condition_Column)
    insert_to_excel(qty[k], new_row, qty_Column)
    insert_to_excel(item_amount[k], new_row, item_amount_Column)
    insert_to_excel(unit[k], new_row, unit_Column)
    insert_to_excel(reported_cost[k], new_row, reported_cost_Column)
    insert_to_excel(unit_cost[k], new_row, unit_cost_Column)
    insert_to_excel(coverage[k], new_row, coverage_Column)
    insert_to_excel(sales_tax[k], new_row, sales_tax_Column)
    insert_to_excel(rcv[k], new_row, rcv_Column)
    insert_to_excel(life[k], new_row, life_Column)
    insert_to_excel(depreciation_type[k], new_row, depreciation_type_Column)
    insert_to_excel(depreciation_amount[k], new_row, depreciation_amount_Column)
    insert_to_excel(recoverable[k], new_row, recoverable_Column)
    insert_to_excel(acv[k], new_row, acv_Column)
    insert_to_excel(tax[k], new_row, tax_Column)
    insert_to_excel(replace[k], new_row, replace_Column)
    insert_to_excel(cat[k], new_row, cat_Column)
    insert_to_excel(sel[k], new_row, sel_Column)
    insert_to_excel(owner[k], new_row, owner_Column)
    insert_to_excel(original_vendor[k], new_row, original_vendor_Column)
    insert_to_excel(date[k], new_row, date_Column)
    insert_to_excel(note_1[k], new_row, note_1_Column)
    new_row += 1
    k += 1

start()


# resize columns for more readable
for col in sheet.columns:
     max_length = 0
     column = col[0].column_letter # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     sheet.column_dimensions[column].width = adjusted_width

wb.save(path) 